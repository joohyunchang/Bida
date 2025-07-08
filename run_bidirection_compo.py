import argparse
import datetime
import deepspeed
import numpy as np
import time
import torch
import torch.nn as nn
import torch.backends.cudnn as cudnn
import json
import os
from functools import partial
from pathlib import Path
from collections import OrderedDict

from util_tools.mixup import Mixup
from timm.models import create_model
from timm.loss import LabelSmoothingCrossEntropy, SoftTargetCrossEntropy
from timm.utils import ModelEma
from util_tools.optim_factory import create_optimizer, get_parameter_groups, LayerDecayValueAssigner

from dataset.datasets import build_dataset
from util_tools.utils import NativeScalerWithGradNormCount as NativeScaler, load_bidir_weights, unfreeze_block, freeze_block_list
from util_tools.utils import multiple_samples_collate, notice_message, laod_eval_weights, audio_collate_fn, test_audio_collate_fn, audio_list_collate_fn, test_audio_list_collate_fn
import util_tools.utils as utils
import models.bidir_modeling_after_crossattn
import models.bidir_modeling_crossattn
import models.bidir_aim_modeling_crossattn
import models.bidir_modeling_crossattn_concat
import models.cast_square
import models.cast_bisquare
import models.cast_Bsquare
import models.ast_Bsquare
import models.AIM_cls
import models.audio_cast
import models.audio_only
import models.audio_clip_cast
import models.ast_clip_cast
import models.ast_vmae_cast
import models.beats_clip_cast
import models.beats_Bsquare
from models.prompt import text_prompt, dataset_class
import pandas as pd
from timm.models.registry import register_model
from transformers import VideoMAEForVideoClassification, CLIPModel
# from models.paraphrase import paraphrase


def get_args():
    parser = argparse.ArgumentParser('VideoMAE fine-tuning and evaluation script for video classification', add_help=False)
    parser.add_argument('--batch_size', default=64, type=int)
    parser.add_argument('--epochs', default=30, type=int)
    parser.add_argument('--update_freq', default=1, type=int)
    parser.add_argument('--save_ckpt_freq', default=100, type=int)
    parser.add_argument('--debug', action='store_true', default=False)

    # Model parameters
    parser.add_argument('--vmae_model', default='vit_base_patch16_224', type=str, metavar='MODEL',
                        help='Name of model to train')
    parser.add_argument('--clip_frame', default=None, type=str)
    parser.add_argument('--tubelet_size', type=int, default= 2)
    parser.add_argument('--input_size', default=224, type=int,
                        help='videos input size')

    parser.add_argument('--drop', type=float, default=0.0, metavar='PCT',
                        help='Dropout rate (default: 0.)')
    parser.add_argument('--head_drop', type=float, default=0.0, metavar='PCT',
                        help='Dropout rate (default: 0.)')
    parser.add_argument('--attn_drop_rate', type=float, default=0.0, metavar='PCT',
                        help='Attention dropout rate (default: 0.)')
    parser.add_argument('--drop_path', type=float, default=0.1, metavar='PCT',
                        help='Drop path rate (default: 0.1)')

    parser.add_argument('--disable_eval_during_finetuning', action='store_true', default=False)
    parser.add_argument('--model_ema', action='store_true', default=False)
    parser.add_argument('--model_ema_decay', type=float, default=0.9999, help='')
    parser.add_argument('--model_ema_force_cpu', action='store_true', default=False, help='')

    # Optimizer parameters
    parser.add_argument('--focal_loss_gamma', default=None, type=float,
                        help='focal loss gamma')
    parser.add_argument('--opt', default='adamw', type=str, metavar='OPTIMIZER',
                        help='Optimizer (default: "adamw"')
    parser.add_argument('--opt_eps', default=1e-8, type=float, metavar='EPSILON',
                        help='Optimizer Epsilon (default: 1e-8)')
    parser.add_argument('--opt_betas', default=None, type=float, nargs='+', metavar='BETA',
                        help='Optimizer Betas (default: None, use opt default)')
    parser.add_argument('--clip_grad', type=float, default=None, metavar='NORM',
                        help='Clip gradient norm (default: None, no clipping)')
    parser.add_argument('--momentum', type=float, default=0.9, metavar='M',
                        help='SGD momentum (default: 0.9)')
    parser.add_argument('--weight_decay', type=float, default=0.05,
                        help='weight decay (default: 0.05)')
    parser.add_argument('--weight_decay_end', type=float, default=None, help="""Final value of the
        weight decay. We use a cosine schedule for WD and using a larger decay by
        the end of training improves performance for ViTs.""")

    parser.add_argument('--lr', type=float, default=1e-3, metavar='LR',
                        help='learning rate (default: 1e-3)')
    parser.add_argument('--layer_decay', type=float, default=0.75)

    parser.add_argument('--warmup_lr', type=float, default=1e-6, metavar='LR',
                        help='warmup learning rate (default: 1e-6)')
    parser.add_argument('--min_lr', type=float, default=1e-6, metavar='LR',
                        help='lower lr bound for cyclic schedulers that hit 0 (1e-5)')

    parser.add_argument('--warmup_epochs', type=int, default=5, metavar='N',
                        help='epochs to warmup LR, if scheduler supports')
    parser.add_argument('--warmup_steps', type=int, default=-1, metavar='N',
                        help='num of steps to warmup LR, will overload warmup_epochs if set > 0')

    # Augmentation parameters
    parser.add_argument('--color_jitter', type=float, default=0.4, metavar='PCT',
                        help='Color jitter factor (default: 0.4)')
    parser.add_argument('--num_sample', type=int, default=2,
                        help='Repeated_aug (default: 2)')
    parser.add_argument('--aa', type=str, default='rand-m7-n4-mstd0.5-inc1', metavar='NAME',
                        help='Use AutoAugment policy. "v0" or "original". " + "(default: rand-m7-n4-mstd0.5-inc1)'),
    parser.add_argument('--smoothing', type=float, default=0.1,
                        help='Label smoothing (default: 0.1)')
    parser.add_argument('--train_interpolation', type=str, default='bicubic',
                        help='Training interpolation (random, bilinear, bicubic default: "bicubic")')

    # Evaluation parameters
    parser.add_argument('--crop_pct', type=float, default=None)
    parser.add_argument('--short_side_size', type=int, default=224)
    parser.add_argument('--test_num_segment', type=int, default=5)
    parser.add_argument('--test_num_crop', type=int, default=3)
    
    # Random Erase params
    parser.add_argument('--reprob', type=float, default=0.25, metavar='PCT',
                        help='Random erase prob (default: 0.25)')
    parser.add_argument('--remode', type=str, default='pixel',
                        help='Random erase mode (default: "pixel")')
    parser.add_argument('--recount', type=int, default=1,
                        help='Random erase count (default: 1)')
    parser.add_argument('--resplit', action='store_true', default=False,
                        help='Do not random erase first (clean) augmentation split')

    # Mixup params
    parser.add_argument('--mixup', type=float, default=0.8,
                        help='mixup alpha, mixup enabled if > 0.')
    parser.add_argument('--cutmix', type=float, default=1.0,
                        help='cutmix alpha, cutmix enabled if > 0.')
    parser.add_argument('--cutmix_minmax', type=float, nargs='+', default=None,
                        help='cutmix min/max ratio, overrides alpha and enables cutmix if set (default: None)')
    parser.add_argument('--mixup_prob', type=float, default=1.0,
                        help='Probability of performing mixup or cutmix when either/both is enabled')
    parser.add_argument('--mixup_switch_prob', type=float, default=0.5,
                        help='Probability of switching to cutmix when both mixup and cutmix enabled')
    parser.add_argument('--mixup_mode', type=str, default='batch',
                        help='How to apply mixup/cutmix params. Per "batch", "pair", or "elem"')

    # Finetuning params
    parser.add_argument('--vmae_finetune', default='', help='finetune from checkpoint')
    parser.add_argument('--clip_finetune',default='', help='finetune from clip checkpoint')
    parser.add_argument('--fine_tune', default=None, help='finetune from bidir model')
    parser.add_argument('--model_key', default='model|module', type=str)
    parser.add_argument('--model_prefix', default='', type=str)
    parser.add_argument('--init_scale', default=0.001, type=float)
    parser.add_argument('--use_mean_pooling', action='store_true')
    parser.set_defaults(use_mean_pooling=True)
    parser.add_argument('--use_cls', action='store_false', dest='use_mean_pooling')

    # Dataset parameters
    parser.add_argument('--data_path', default='/path/to/list_kinetics-400', type=str,
                        help='dataset path')
    parser.add_argument('--anno_path', default=None, type=str, help='annotation path')
    parser.add_argument('--eval_data_path', default=None, type=str,
                        help='dataset path for evaluation')
    parser.add_argument('--nb_classes', default=400, type=int,
                        help='number of the classification types')
    parser.add_argument('--imagenet_default_mean_and_std', default=True, action='store_true')
    parser.add_argument('--num_segments', type=int, default= 1)
    parser.add_argument('--num_frames', type=int, default= 16)
    parser.add_argument('--sampling_rate', type=int, default= 4)
    parser.add_argument('--data_set', default='Kinetics-400', choices=['ActivityNet','HD_EPIC','EPIC_dense','diving-48','Kinetics-400', 'SSV2','MINI_SSV2', 'UCF101', 'HMDB51','image_folder', 'EPIC','Kinetics_sound', 'EPIC_sounds','VGGSound'],
                        type=str, help='dataset')
    parser.add_argument('--pred_type', default=None, choices=['noun', 'verb', 'action'])
    parser.add_argument('--output_dir', default='',
                        help='path where to save, empty for no saving')
    parser.add_argument('--log_dir', default=None,
                        help='path where to tensorboard log')
    parser.add_argument('--device', default='cuda',
                        help='device to use for training / testing')
    parser.add_argument('--seed', default=0, type=int)
    parser.add_argument('--resume', default='',
                        help='resume from checkpoint')
    parser.add_argument('--auto_resume', action='store_true')
    parser.add_argument('--no_auto_resume', action='store_false', dest='auto_resume')
    parser.set_defaults(auto_resume=True)

    parser.add_argument('--save_ckpt', action='store_true')
    parser.add_argument('--no_save_ckpt', action='store_false', dest='save_ckpt')
    parser.set_defaults(save_ckpt=True)

    parser.add_argument('--start_epoch', default=0, type=int, metavar='N',
                        help='start epoch')
    parser.add_argument('--eval', action='store_true',
                        help='Perform evaluation only')
    parser.add_argument('--dist_eval', action='store_true', default=False,
                        help='Enabling distributed evaluation')
    parser.add_argument('--num_workers', default=10, type=int)
    parser.add_argument('--pin_mem', action='store_true',
                        help='Pin CPU memory in DataLoader for more efficient (sometimes) transfer to GPU.')
    parser.add_argument('--no_pin_mem', action='store_false', dest='pin_mem')
    parser.set_defaults(pin_mem=True)

    # distributed training parameters
    parser.add_argument('--world_size', default=1, type=int,
                        help='number of distributed processes')
    parser.add_argument('--local_rank', type=int)
    parser.add_argument('--local-rank', type=int)
    parser.add_argument('--dist_on_itp', action='store_true')
    parser.add_argument('--dist_url', default='env://',
                        help='url used to set up distributed training')

    parser.add_argument('--enable_deepspeed', action='store_true', default=False)
    # new settings
    parser.add_argument('--freeze_layers', default=None, nargs='+', type=str)
    parser.add_argument('--unfreeze_layers', default=None, nargs='+', type=str)
    parser.add_argument('--slack_api', type=str,default=None)
    parser.add_argument('--composition', action='store_true')
    parser.add_argument('--fusion_method', default='add', choices=['add','mul','concat','weight'],
                        type=str, help='fusion_method')
    parser.add_argument('--throughput', action='store_true', dest='throughput')
    parser.add_argument('--eval_result', action='store_true', default=False,
                        help='Perform evaluation only')
    parser.add_argument('--xlsx', action='store_true', default=False)
    parser.add_argument('--text_finetune',default=None, help='finetune from clip checkpoint')
    parser.add_argument('--audio_finetune',default=None, help='finetune from clip checkpoint')
    parser.add_argument('--prompt_weight',default=None, help='prompt from prompt_cast checkpoint')
    parser.add_argument('--audio_path', default=None, type=str, help='audio path')
    parser.add_argument('--collate', action='store_true', default=False)
    parser.add_argument('--audio_type', default='all8', choices=['all','all8','frame','stack','stacks','single','onespec','single1024','stackss','single1024s','singles','beats_single128','beats_single224','beats_single1024','beats_free','free'],
                        type=str, help='audio_trim_type')
    parser.add_argument('--narration', action='store_true', default=False)
    parser.add_argument('--class_narration', action='store_true', default=False)
    parser.add_argument('--spec_augment', action='store_true', default=False)
    parser.add_argument('--realtime_audio', action='store_true', default=False)
    parser.add_argument('--audio_height', default=224, type=int, help='audio_spec_shape')
    parser.add_argument('--audio_width', default=224, type=int, help='audio_spec_shape')
    parser.add_argument('--autosave_spec', action='store_true', default=False)
    parser.add_argument('--noisereduce', action='store_true', default=False)
    parser.add_argument('--specnorm', action='store_true', default=False)
    parser.add_argument('--bcast_method', default=None, choices=['seq','add','add_scale','add_param','msa_add'], # sequential, parallel add, parallel add scale
                        type=str, help='bcast_method')
    parser.add_argument('--process_type', type=str,default='ast')
    parser.add_argument('--ucf101_type', type=str, default='1')
    parser.add_argument('--time_encoding', action='store_true', default=False)
    parser.add_argument('--disable_video', action='store_true', default=False)
    parser.add_argument('--audio_only_finetune',default=None, help='finetune from clip checkpoint')
    parser.add_argument('--ast_finetune',default=None, help='finetune from clip checkpoint')
    parser.add_argument('--stride', type=int, default=10)
    parser.add_argument('--enable_audio_stride', action='store_true', default=False)
    parser.add_argument('--mixup_spec', action='store_true', default=False)
    parser.add_argument('--spec_cutmix', action='store_true', default=False)
    parser.add_argument('--add_noise', action='store_true', default=False)
    parser.add_argument('--not_use_stpos', action='store_false', default=True)
    parser.add_argument('--pre_time_encoding', action='store_true', default=False)
    parser.add_argument('--split_time_mlp', action='store_true', default=False)
    parser.add_argument('--bcast_share', action='store_true', default=False)
    parser.add_argument('--imagenet', default=None, help='finetune from clip imagenet checkpoint')
    parser.add_argument('--fixpatch', action='store_true', default=False)
    parser.add_argument('--videomae_v2', action='store_true', default=False, help='finetune from VideMAE V2 checkpoint')
    parser.add_argument('--ablation_eval', default=None, choices=['white_noise','pink_noise','missing','time_shift','msa_add'],
                        type=str, help='ablation_eval')
    parser.add_argument('--disable_load_weights', action='store_true', default=False)
    
    class VideoCLIPWithHead(nn.Module):
        def __init__(self, num_classes=400, feature_agg='mean'):
            super().__init__()
            # load CLIP vision backbone
            model = CLIPModel.from_pretrained('openai/clip-vit-base-patch16', use_safetensors=True)
            self.clip_vision = model.vision_model
            self.head = nn.Linear(self.clip_vision.config.hidden_size, num_classes)
            self.feature_agg = feature_agg  # 'mean' or 'sum'
        
        def get_num_layers(self):
            return len(self.clip_vision.encoder.layers)
        
        @torch.jit.ignore
        def no_weight_decay(self):
            return {'embeddings'}

        def forward(self, x, **kwargs):  # x: [B, C, T, H, W]
            B, C, T, H, W = x.shape
            # (1) Frame 단위로 CLIP backbone 통과 (batch concat)
            x = x.permute(0, 2, 1, 3, 4).reshape(B * T, C, H, W)  # [B*T, C, H, W]
            outputs = self.clip_vision(pixel_values=x, return_dict=True)
            features = outputs.pooler_output  # [B*T, D]
            features = features.view(B, T, -1)  # [B, T, D]
            # (2) 모든 frame feature를 sum/mean pooling
            if self.feature_agg == 'mean':
                vid_feat = features.mean(dim=1)
            elif self.feature_agg == 'sum':
                vid_feat = features.sum(dim=1)
            else:
                raise NotImplementedError
            out = self.head(vid_feat)  # [B, num_classes]
            return out
        
    class VideoMAEWithHead(nn.Module):
        def __init__(self, num_classes=400):
            super().__init__()
            self.model = VideoMAEForVideoClassification.from_pretrained(
                'MCG-NJU/videomae-base-finetuned-kinetics')
            self.model.classifier = torch.nn.Linear(self.model.classifier.in_features, num_classes)
            
        def get_num_layers(self):
            return len(self.model.videomae.encoder.layer)
        
        @torch.jit.ignore
        def no_weight_decay(self):
            return {'embeddings'}

        def forward(self, x, **kwargs):  # x: [B, C, T, H, W]
            outputs = self.model(pixel_values=x.permute(0, 2, 1, 3, 4), return_dict=True) # [B, num_classes]
            out = outputs.logits  # [B, num_classes]
            return out
    
    @register_model
    def clip_model(pretrained=False, **kwargs):
        print('num_classes = %s' % kwargs.get('num_classes', None))
        model = VideoCLIPWithHead(num_classes=kwargs.get('num_classes', 400), feature_agg='mean')
        return model
    
    @register_model
    def videomae_v1_model(pretrained=False, **kwargs):
        model = VideoMAEWithHead(num_classes=kwargs.get('num_classes', 400))
        return model

    known_args, _ = parser.parse_known_args()

    if known_args.enable_deepspeed:
        try:
            parser = deepspeed.add_config_arguments(parser)
            ds_init = deepspeed.initialize
        except:
            print("Please 'pip install deepspeed'")
            exit(0)
    else:
        ds_init = None

    return parser.parse_args(), ds_init

def parse_tuple(string):
    try:
        return tuple(map(int, string.split(',')))
    except:
        raise argparse.ArgumentTypeError("Tuple must be x,y")

def main(args, ds_init):
    utils.init_distributed_mode(args)

    if ds_init is not None:
        utils.create_ds_config(args)

    print(args)

    device = torch.device(args.device)

    # fix the seed for reproducibility
    seed = args.seed + utils.get_rank()
    torch.manual_seed(seed)
    np.random.seed(seed)
    # random.seed(seed)

    cudnn.benchmark = True
    # args.process_type = 'beats' if 'beats' in args.vmae_model else args.process_type
    args.audio_path = None if 'all' in args.ucf101_type else args.audio_path
    
    dataset_train, args.nb_classes = build_dataset(is_train=True, test_mode=False, args=args)
    if args.disable_eval_during_finetuning:
        dataset_val = None
    else:
        dataset_val, _ = build_dataset(is_train=False, test_mode=False, args=args)
    dataset_test, _ = build_dataset(is_train=False, test_mode=True, args=args)
    
    if args.ast_finetune is not None and 'SSAST' in args.ast_finetune:
        args.stride = 16

    num_tasks = utils.get_world_size()
    global_rank = utils.get_rank()
    sampler_train = torch.utils.data.DistributedSampler(
        dataset_train, num_replicas=num_tasks, rank=global_rank, shuffle=True
    )
    print("Sampler_train = %s" % str(sampler_train))
    if args.dist_eval:
        if len(dataset_val) % num_tasks != 0:
            print('Warning: Enabling distributed evaluation with an eval dataset not divisible by process number. '
                    'This will slightly alter validation results as extra duplicate entries are added to achieve '
                    'equal num of samples per-process.')
        sampler_val = torch.utils.data.DistributedSampler(
            dataset_val, num_replicas=num_tasks, rank=global_rank, shuffle=False)
        sampler_test = torch.utils.data.DistributedSampler(
            dataset_test, num_replicas=num_tasks, rank=global_rank, shuffle=False)
    else:
        sampler_val = torch.utils.data.SequentialSampler(dataset_val)
        sampler_test = torch.utils.data.SequentialSampler(dataset_test)

    if global_rank == 0 and args.log_dir is not None:
        os.makedirs(args.log_dir, exist_ok=True)
        log_writer = utils.TensorboardLogger(log_dir=args.log_dir)
    else:
        log_writer = None

    # if args.num_sample > 1:
    #     collate_func = partial(multiple_samples_collate, fold=False)
    # else:
    #     collate_func = None
    if args.collate:
        train_collate, val_collate, test_collate = audio_list_collate_fn, audio_list_collate_fn, test_audio_list_collate_fn
    elif args.audio_path is not None or args.narration is not None or args.class_narration is not None:
        train_collate, val_collate, test_collate = audio_collate_fn, audio_collate_fn, test_audio_collate_fn
    else:
        train_collate, val_collate, test_collate = None, None, None

    data_loader_train = torch.utils.data.DataLoader(
        dataset_train, sampler=sampler_train,
        batch_size=args.batch_size,
        num_workers=args.num_workers,
        pin_memory=args.pin_mem,
        drop_last=True,
        collate_fn=train_collate,
    )

    if dataset_val is not None:
        data_loader_val = torch.utils.data.DataLoader(
            dataset_val, sampler=sampler_val,
            batch_size=int(1.5 * args.batch_size),
            num_workers=args.num_workers,
            pin_memory=args.pin_mem,
            drop_last=False,
            collate_fn=val_collate
        )
    else:
        data_loader_val = None

    if dataset_test is not None:
        data_loader_test = torch.utils.data.DataLoader(
            dataset_test, sampler=sampler_test,
            batch_size=args.batch_size,
            num_workers=args.num_workers,
            pin_memory=args.pin_mem,
            drop_last=False,
            collate_fn=test_collate
        )
    else:
        data_loader_test = None

    mixup_fn = None
    mixup_active = args.mixup > 0 or args.cutmix > 0. or args.cutmix_minmax is not None
    if mixup_active:
        print("Mixup is activated!")
        mixup_fn = Mixup(
            mixup_alpha=args.mixup, cutmix_alpha=args.cutmix, cutmix_minmax=args.cutmix_minmax,
            prob=args.mixup_prob, switch_prob=args.mixup_switch_prob, mode=args.mixup_mode,
            label_smoothing=args.smoothing, num_classes=args.nb_classes, composition=args.composition, spec_cutmix=args.spec_cutmix)

    patch_size = 14
    print("Patch size = %s" % str(patch_size))
    args.window_size = 16
    args.patch_size = patch_size
    
    
    if False:
        class_list = text_prompt(dataset=args.data_set, data_path=args.anno_path, clipbackbone=args.clip_finetune, device=device)
    else:
        class_list = None
    model_args = {
            'model_name': args.vmae_model,
            'pretrained': False,
            'num_classes': args.nb_classes,
            'all_frames': args.num_frames * args.num_segments,
            'tubelet_size': args.tubelet_size,
            'drop_rate': args.drop,
            'drop_path_rate': args.drop_path,
            'attn_drop_rate': args.attn_drop_rate,
            'drop_block_rate': None,
            'use_mean_pooling': args.use_mean_pooling,
            'init_scale': args.init_scale,
            'fusion_method': args.fusion_method
        }
    # args.fixpatch = True
    if args.audio_path is not None:
        def get_shapes(fstride, tstride, input_fdim=128, input_tdim=1024, fshape=16, tshape=16):
            test_input = torch.randn(1, 1, input_fdim, input_tdim)
            test_proj = nn.Conv2d(1, 768, kernel_size=(fshape, tshape), stride=(fstride, tstride))
            test_out = test_proj(test_input)
            f_dim = test_out.shape[2]
            t_dim = test_out.shape[3]
            return f_dim, t_dim
        if args.fixpatch:
            f_dim, t_dim = get_shapes(args.stride, args.stride, args.audio_height, args.audio_width)
            model_args['audio_patch'] = f_dim * t_dim
            model_args['spec_shape'] = [f_dim, t_dim]
        else:
            f_dim, t_dim = args.audio_height//args.window_size, args.audio_width// args.window_size
            model_args['audio_patch'] = f_dim * t_dim
            model_args['spec_shape'] = [f_dim, t_dim]
        print(f'stride {args.stride}, audio shape: {f_dim}, {t_dim}, patch : {f_dim * t_dim}')
        # print(f"Audio_Patch size = {args.audio_height*args.audio_width//(args.window_size*args.window_size)}")
    if args.bcast_method is not None:
        print(f"bcast_method = {args.bcast_method}")
        model_args['bcast_method'] = args.bcast_method
    if args.time_encoding:
        model_args['time_encoding'] = args.time_encoding
        # model_args['spec_shape'] = [args.audio_height//args.window_size, args.audio_width//args.window_size]
    if args.audio_only_finetune:
        model_args['audio_only_finetune'] = True
    if '_ast_' in args.vmae_model:
        model_args['fstride'] = args.stride
        model_args['tstride'] = args.stride
        model_args['input_fdim'] = args.audio_height
        model_args['input_tdim'] = args.audio_width
    if args.enable_audio_stride:
        model_args['fstride'] = 10
        model_args['tstride'] = 10
        fdim, tdim = int((args.audio_height-16)/10)+1, int((args.audio_width-16)/10)+1
        model_args['audio_patch'] = fdim * tdim
        model_args['spec_shape'] = [fdim, tdim]
    if args.not_use_stpos == False:
        model_args['use_stpos'] = args.not_use_stpos
    if args.pre_time_encoding == True:
        model_args['pre_time_encoding'] = args.pre_time_encoding
    if args.split_time_mlp == True:
        model_args['split_time_mlp'] = args.split_time_mlp
    if args.bcast_share == True:
        model_args['bcast_share'] = args.bcast_share
        
    model = create_model(**model_args)
    before_n_parameters = sum(p.numel() for p in model.parameters() if p.requires_grad)
    print('Before Freeze number of params:', before_n_parameters)
    
    if not getattr(args,'vmae_model', False) in ['videomae_v1_model', 'clip_model']:
        freeze_list = freeze_block_list(model,args.unfreeze_layers)
    if getattr(args,'disable_load_weights', False) is True:
        print("Disable load weights")
    elif args.fine_tune is not None:
        laod_eval_weights(model, args.fine_tune, args)
    else:
        load_bidir_weights(model, args, freeze_list=freeze_list)
    
    # if args.audio_only_finetune is not None:
    #     audio_key = torch.load(args.audio_only_finetune, map_location='cpu')['module']
    #     for k in ['head.weight', 'head.bias', 'head_noun.weight', 'head_noun.bias', 'head_verb.weight', 'head_verb.bias', 'audio_ln_post.weight' 'audio_ln_post.bias']:
    #         if k in audio_key:
    #             del audio_key[k]
    #     modified_data = OrderedDict()
    #     for key, value in audio_key.items():
    #         new_key = key.replace('clip', 'audio')
    #         modified_data[new_key] = value
    #     audio_key = modified_data
    #     model.load_state_dict(audio_key, strict=False)
        
    ###### VMAE 검증을 위해 freeze는 잠시 꺼둔다 #############
    if getattr(args,'vmae_model', False) in ['videomae_v1_model', 'clip_model']:
        for param in model.parameters():
            param.requires_grad = True
    elif args.unfreeze_layers is not None:
        model, unfreeze_list = unfreeze_block(model,args.unfreeze_layers)
        print('unfreeze list :', unfreeze_list)
    # with torch.no_grad():#! module_layers에 들어가는 것만 한다. 
    #     model.clip_patch_embed.weight.copy_(model.patch_embed.proj.weight)
    #     model.clip_patch_embed.bias.copy_(model.patch_embed.proj.bias)
    #     print("patch_embed_initialize")

    
    model.to(device)
    
    
    model_ema = None
    if args.model_ema:
        model_ema = ModelEma(
            model,
            decay=args.model_ema_decay,
            device='cpu' if args.model_ema_force_cpu else '',
            resume='')
        print("Using EMA with decay = %.8f" % args.model_ema_decay)

    model_without_ddp = model
    n_parameters = sum(p.numel() for p in model.parameters() if p.requires_grad)

    print("Model = %s" % str(model_without_ddp))
    print('number of params:', n_parameters)

    total_batch_size = args.batch_size * args.update_freq * utils.get_world_size()
    print(args.batch_size, args.update_freq, utils.get_world_size())
    num_training_steps_per_epoch = len(dataset_train) // total_batch_size
    args.lr = args.lr * total_batch_size / 256
    args.min_lr = args.min_lr * total_batch_size / 256
    args.warmup_lr = args.warmup_lr * total_batch_size / 256
    print("LR = %.8f" % args.lr)
    print("Batch size = %d" % total_batch_size)
    print("Update frequent = %d" % args.update_freq)
    print("Number of training examples = %d" % len(dataset_train))
    print("Number of training training per epoch = %d" % num_training_steps_per_epoch)

    num_layers = model_without_ddp.get_num_layers()
    if args.layer_decay < 1.0:
        assigner = LayerDecayValueAssigner(list(args.layer_decay ** (num_layers + 1 - i) for i in range(num_layers + 2)))
    else:
        assigner = None

    if assigner is not None:
        print("Assigned values = %s" % str(assigner.values))

    skip_weight_decay_list = model.no_weight_decay()
    print("Skip weight decay list: ", skip_weight_decay_list)

    if args.enable_deepspeed:
        loss_scaler = None
        optimizer_params = get_parameter_groups(
            model, args.weight_decay, skip_weight_decay_list,
            assigner.get_layer_id if assigner is not None else None,
            assigner.get_scale if assigner is not None else None)
        model, optimizer, _, _ = ds_init(
            args=args, model=model, model_parameters=optimizer_params, dist_init_required=not args.distributed,
        )

        print("model.gradient_accumulation_steps() = %d" % model.gradient_accumulation_steps())
        assert model.gradient_accumulation_steps() == args.update_freq
    else:
        if args.distributed:
            model = torch.nn.parallel.DistributedDataParallel(model, device_ids=[args.gpu], find_unused_parameters=True)
            model_without_ddp = model.module

        optimizer = create_optimizer(
            args, model_without_ddp, skip_list=skip_weight_decay_list,
            get_num_layer=assigner.get_layer_id if assigner is not None else None, 
            get_layer_scale=assigner.get_scale if assigner is not None else None)
        loss_scaler = NativeScaler()

    print("Use step level LR scheduler!")
    lr_schedule_values = utils.cosine_scheduler(
        args.lr, args.min_lr, args.epochs, num_training_steps_per_epoch,
        warmup_epochs=args.warmup_epochs, warmup_steps=args.warmup_steps,
    )
    if args.weight_decay_end is None:
        args.weight_decay_end = args.weight_decay
    wd_schedule_values = utils.cosine_scheduler(
        args.weight_decay, args.weight_decay_end, args.epochs, num_training_steps_per_epoch)
    print("Max WD = %.7f, Min WD = %.7f" % (max(wd_schedule_values), min(wd_schedule_values)))

    if mixup_fn is not None:
        # smoothing is handled with mixup label transform
        criterion = SoftTargetCrossEntropy()
    elif args.smoothing > 0.:
        criterion = LabelSmoothingCrossEntropy(smoothing=args.smoothing)
    else:
        criterion = torch.nn.CrossEntropyLoss()

    print("criterion = %s" % str(criterion))
    print('number of params:', n_parameters)
    if args.audio_path is not None:
        print(f'stride {args.stride}, audio shape: {f_dim}, {t_dim}, patch : {f_dim * t_dim}')
    print("NoiseReduce On") if args.noisereduce else print("NoiseReduce Off")
    
    utils.auto_load_model(
        args=args, model=model, model_without_ddp=model_without_ddp,
        optimizer=optimizer, loss_scaler=loss_scaler, model_ema=model_ema)
    
    if args.composition:
        from engine_for_compomodel import train_one_epoch, validation_one_epoch, final_test, merge
    else:
        from engine_for_finetuning import train_one_epoch, validation_one_epoch, final_test, merge, speedup_one_epoch
    
    if args.eval or args.eval_result:
        if args.throughput:
            avg=speedup_one_epoch(args, data_loader_test,model,device)
            print('ave_forward_throughput is {:.4f}'.format(avg))
            # print('total_throughput is {:.4f}'.format(total))
            exit(0)
        else:
            preds_file = os.path.join(args.output_dir, str(global_rank) + '.txt')
            if not args.eval_result:
                test_stats = final_test(args, data_loader_test, model, device, preds_file)
            torch.distributed.barrier()
            # class_list = text_prompt(dataset=args.data_set, data_path=args.anno_path, clipbackbone=args.clip_finetune, device=device)
            class_list = dataset_class(dataset=args.data_set, data_path=args.anno_path)
            if global_rank == 0:
                print("Start merging results...")
                if args.composition:
                    final_top1_action ,final_top5_action, final_top1_noun, final_top5_noun, final_top1_verb, final_top5_verb, pred_noun, pred_verb, label_noun, label_verb, video_ids, conf_noun, conf_verb = merge(args.output_dir, num_tasks, return_result=True)
                    print(f"Accuracy of the network on the {len(dataset_test)} test videos: Top-1: {final_top1_action:.2f}%, Top-5: {final_top5_action:.2f}%")
                    log_stats = {'Final Top-1 Action': final_top1_action,
                                'Final Top-5 Action': final_top5_action,
                                'Final Top-1 Noun': final_top1_noun,
                                'Final Top-1 Verb': final_top1_verb,
                                'Final Top-5 Noun': final_top5_noun,
                                'Final Top-5 Verb': final_top5_verb,
                                'confidences_noun': np.array(conf_noun).mean(),
                                'confidences_verb': np.array(conf_verb).mean()}
                    
                    # ======== save prediction result ======== #
                    video_ids = [''.join(x).replace(' ','') for x in video_ids]
                    pred_noun = [class_list['noun'][i] for i in pred_noun]
                    pred_verb = [class_list['verb'][i] for i in pred_verb]
                    label_noun = [class_list['noun'][int(i)] for i in label_noun]
                    label_verb = [class_list['verb'][int(i)] for i in label_verb]
                    pred_df = pd.DataFrame({'video_id':video_ids, 'verb':pred_verb, 'noun':pred_noun, 'label_verb':label_verb, 'label_noun':label_noun, 'conf_verb':conf_verb, 'conf_noun':conf_noun})
                    pred_df['action'] = pred_df['verb'] + ' ' + pred_df['noun']
                    pred_df['conf_verb'], pred_df['conf_noun'] = (pred_df['conf_verb']).round(4), (pred_df['conf_noun']).round(4)
                    pred_df.to_csv(os.path.join(args.output_dir + "/../", 'pred_result.csv'), index=False)
                    
                    if args.xlsx:
                        from openpyxl import Workbook
                        from openpyxl.styles import PatternFill
                        from openpyxl.utils.dataframe import dataframe_to_rows

                        wb = Workbook()
                        ws = wb.active

                        # 데이터프레임을 엑셀 시트로 변환
                        for r in dataframe_to_rows(pred_df, index=False, header=True):
                            ws.append(r) 
                            
                        red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
                        for row in ws.iter_rows(min_row=2, max_col=7, max_row=len(pred_df) + 1):
                            if row[1].value != row[3].value:
                                row[1].fill = red_fill
                                row[5].fill = red_fill
                            if row[2].value != row[4].value:
                                row[2].fill = red_fill
                                row[6].fill = red_fill
                        wb.save(os.path.join(args.output_dir + "/../", 'pred_result.xlsx'))
                else:
                    print("Start merging results...")
                    final_top1 ,final_top5, pred, label, video_ids, conf = merge(args.output_dir, num_tasks, return_result=True)
                    print(f"Accuracy of the network on the {len(dataset_test)} test videos: Top-1: {final_top1:.2f}%, Top-5: {final_top5:.2f}%")
                    log_stats = {'Final top-1': final_top1, 
                                'Final Top-5': final_top5,
                                'confidences': np.array(conf).mean()}
                    
                    # ======== save prediction result ======== #
                    video_ids = [''.join(x).replace(' ','') for x in video_ids]
                    pred = [class_list['action'][i] for i in pred]
                    label = [class_list['action'][int(i)] for i in label]
                    pred_df = pd.DataFrame({'video_id':video_ids, 'pred':pred, 'label':label, 'conf':conf})
                    pred_df['conf'] = (pred_df['conf']).round(4)
                    pred_df.to_csv(os.path.join(args.output_dir + "/../", 'pred_result.csv'), index=False)
                    
                    if args.xlsx:
                        from openpyxl import Workbook
                        from openpyxl.styles import PatternFill
                        from openpyxl.utils.dataframe import dataframe_to_rows

                        wb = Workbook()
                        ws = wb.active

                        # 데이터프레임을 엑셀 시트로 변환
                        for r in dataframe_to_rows(pred_df, index=False, header=True):
                            ws.append(r) 

                        red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
                        for row in ws.iter_rows(min_row=2, max_col=4, max_row=len(pred_df) + 1):
                            if row[1].value != row[2].value:
                                row[1].fill = red_fill
                                row[3].fill = red_fill
                        wb.save(os.path.join(args.output_dir + "/../", 'pred_result.xlsx'))
                if args.output_dir and utils.is_main_process():
                    with open(os.path.join(args.output_dir + "/../", "log.txt"), mode="a", encoding="utf-8") as f:
                        f.write(json.dumps(log_stats) + "\n")
            exit(0)
    
    print(f"Start training for {args.epochs} epochs")
    start_time = time.time()
    max_accuracy = 0.0
    torch.cuda.empty_cache()
    epoch = args.start_epoch
    for epoch in range(args.start_epoch, args.epochs):
        if args.distributed:
            data_loader_train.sampler.set_epoch(epoch)
        if log_writer is not None:
            log_writer.set_step(epoch * num_training_steps_per_epoch * args.update_freq)
        train_stats = train_one_epoch(
            args, model, criterion, data_loader_train, optimizer,
            device, epoch, loss_scaler, args.clip_grad, model_ema, mixup_fn,
            log_writer=log_writer, start_steps=epoch * num_training_steps_per_epoch,
            lr_schedule_values=lr_schedule_values, wd_schedule_values=wd_schedule_values,
            num_training_steps_per_epoch=num_training_steps_per_epoch, update_freq=args.update_freq, class_list=class_list
        )
        torch.cuda.empty_cache()
        if args.output_dir and args.save_ckpt:
            if (epoch + 1) % args.save_ckpt_freq == 0 or epoch + 1 == args.epochs:
                utils.save_model(
                    args=args, model=model, model_without_ddp=model_without_ddp, optimizer=optimizer,
                    loss_scaler=loss_scaler, epoch=epoch, model_ema=model_ema)
        if data_loader_val is not None:
            test_stats = validation_one_epoch(args, data_loader_val, model, device)
            torch.cuda.empty_cache()
            
            if args.composition:
                print(f"Accuracy of the network on the {len(dataset_val)} val videos: {test_stats['acc1_action']:.1f}%")
                if max_accuracy < test_stats["acc1_action"]:
                    max_accuracy = test_stats["acc1_action"]
                    if args.output_dir and args.save_ckpt:
                        utils.save_model(
                            args=args, model=model, model_without_ddp=model_without_ddp, optimizer=optimizer,
                            loss_scaler=loss_scaler, epoch="best", model_ema=model_ema, real_epoch=epoch)

                print(f'Max accuracy: {max_accuracy:.2f}%')
                if log_writer is not None:
                    log_writer.update(val_acc1=test_stats['acc1_action'], head="perf", step=epoch)

                log_stats = {**{f'train_{k}': v for k, v in train_stats.items()},
                            **{f'val_{k}': v for k, v in test_stats.items()},
                            'epoch': epoch,
                            'n_parameters': n_parameters}
            else:
                print(f"Accuracy of the network on the {len(dataset_val)} val videos: {test_stats['acc1']:.1f}%")
                if max_accuracy < test_stats["acc1"]:
                    max_accuracy = test_stats["acc1"]
                    if args.output_dir and args.save_ckpt:
                        utils.save_model(
                            args=args, model=model, model_without_ddp=model_without_ddp, optimizer=optimizer,
                            loss_scaler=loss_scaler, epoch="best", model_ema=model_ema, real_epoch=epoch)

                print(f'Max accuracy: {max_accuracy:.2f}%')
                if log_writer is not None:
                    log_writer.update(val_acc1=test_stats['acc1'], head="perf", step=epoch)
                    log_writer.update(val_acc5=test_stats['acc5'], head="perf", step=epoch)
                    log_writer.update(val_loss=test_stats['loss'], head="perf", step=epoch)

                log_stats = {**{f'train_{k}': v for k, v in train_stats.items()},
                            **{f'val_{k}': v for k, v in test_stats.items()},
                            'epoch': epoch,
                            'n_parameters': n_parameters}
        else:
            log_stats = {**{f'train_{k}': v for k, v in train_stats.items()},
                         'epoch': epoch,
                         'n_parameters': n_parameters}
        if args.output_dir and utils.is_main_process():
            if log_writer is not None:
                log_writer.flush()
            with open(os.path.join(args.output_dir + "/../", "log.txt"), mode="a", encoding="utf-8") as f:
                f.write(json.dumps(log_stats) + "\n")
    
    current_epoch = epoch
    acc_str = ""
    for idx in range(2):
        if idx == 1:
            try:
                _, client_states = model.load_checkpoint(args.output_dir, tag='checkpoint-best')
                current_epoch = client_states['epoch']
                if current_epoch == args.epochs - 1:
                    break
                print('Best checkpoint: ', current_epoch)
            except:
                break
            
        preds_file = os.path.join(args.output_dir, str(global_rank) + '.txt')
        test_stats = final_test(args, data_loader_test, model, device, preds_file)
        torch.distributed.barrier()
        if global_rank == 0:
            print("Start merging results...")
            # class_list = text_prompt(dataset=args.data_set, data_path=args.anno_path, clipbackbone=args.clip_finetune, device=device)
            class_list = dataset_class(dataset=args.data_set, data_path=args.anno_path)
            if args.composition:
                final_top1_action ,final_top5_action, final_top1_noun, final_top5_noun, final_top1_verb, final_top5_verb, pred_noun, pred_verb, label_noun, label_verb, video_ids, conf_noun, conf_verb = merge(args.output_dir, num_tasks, return_result=True)
                print(f"Accuracy of the network on the {len(dataset_test)} test videos: Top-1 Action: {final_top1_action:.2f}%, Top-5 Action: {final_top5_action:.2f}%")
                log_stats = {'Epoch': current_epoch,
                            'Final top-1 Action': final_top1_action,
                            'Final Top-5 Action': final_top5_action,
                            'Final Top-1 Noun': final_top1_noun,
                            'Final Top-1 Verb': final_top1_verb,
                            'Final Top-5 Noun': final_top5_noun,
                            'Final Top-5 Verb': final_top5_verb,
                            'confidences_noun': np.array(conf_noun).mean(),
                            'confidences_verb': np.array(conf_verb).mean()}
                if args.output_dir and utils.is_main_process():
                    with open(os.path.join(args.output_dir + "/../", "log.txt"), mode="a", encoding="utf-8") as f:
                        f.write(json.dumps(log_stats) + "\n")
        
                # ======== save prediction result ======== #
                video_ids = [''.join(x).replace(' ','') for x in video_ids]
                pred_noun = [class_list['noun'][i] for i in pred_noun]
                pred_verb = [class_list['verb'][i] for i in pred_verb]
                label_noun = [class_list['noun'][int(i)] for i in label_noun]
                label_verb = [class_list['verb'][int(i)] for i in label_verb]
                pred_df = pd.DataFrame({'video_id':video_ids, 'verb':pred_verb, 'noun':pred_noun, 'label_verb':label_verb, 'label_noun':label_noun, 'conf_verb':conf_verb, 'conf_noun':conf_noun})
                pred_df['action'] = pred_df['verb'] + ' ' + pred_df['noun']
                pred_df['conf_verb'], pred_df['conf_noun'] = (pred_df['conf_verb']).round(4), (pred_df['conf_noun']).round(4)
                pred_df.to_csv(os.path.join(args.output_dir + "/../", 'pred_result.csv'), index=False)
                
                if args.xlsx:
                    from openpyxl import Workbook
                    from openpyxl.styles import PatternFill
                    from openpyxl.utils.dataframe import dataframe_to_rows

                    wb = Workbook()
                    ws = wb.active

                    # 데이터프레임을 엑셀 시트로 변환
                    for r in dataframe_to_rows(pred_df, index=False, header=True):
                        ws.append(r) 
                        
                    red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
                    for row in ws.iter_rows(min_row=2, max_col=7, max_row=len(pred_df) + 1):
                        if row[1].value != row[3].value:
                            row[1].fill = red_fill
                            row[5].fill = red_fill
                        if row[2].value != row[4].value:
                            row[2].fill = red_fill
                            row[6].fill = red_fill
                    wb.save(os.path.join(args.output_dir + "/../", 'pred_result.xlsx'))
            else:
                final_top1 ,final_top5, pred, label, video_ids, conf = merge(args.output_dir, num_tasks, return_result=True)
                print(f"Accuracy of the network on the {len(dataset_test)} test videos: Top-1: {final_top1:.2f}%, Top-5: {final_top5:.2f}%")
                log_stats = {'Epoch': current_epoch,
                            'Final top-1': final_top1, 
                            'Final Top-5': final_top5,
                            'confidences': np.array(conf).mean()}
                if args.output_dir and utils.is_main_process():
                    with open(os.path.join(args.output_dir + "/../", "log.txt"), mode="a", encoding="utf-8") as f:
                        f.write(json.dumps(log_stats) + "\n")
                
                # ======== save prediction result ======== #
                video_ids = [''.join(x).replace(' ','') for x in video_ids]
                pred = [class_list['action'][i] for i in pred]
                label = [class_list['action'][int(i)] for i in label]
                pred_df = pd.DataFrame({'video_id':video_ids, 'pred':pred, 'label':label, 'conf':conf})
                pred_df['conf'] = (pred_df['conf']).round(4)
                pred_df.to_csv(os.path.join(args.output_dir + "/../", 'pred_result.csv'), index=False)
                
                if args.xlsx:
                    from openpyxl import Workbook
                    from openpyxl.styles import PatternFill
                    from openpyxl.utils.dataframe import dataframe_to_rows

                    wb = Workbook()
                    ws = wb.active
                    
                    # 데이터프레임을 엑셀 시트로 변환
                    for r in dataframe_to_rows(pred_df, index=False, header=True):
                        ws.append(r) 

                    red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
                    for row in ws.iter_rows(min_row=2, max_col=4, max_row=len(pred_df) + 1):
                        if row[1].value != row[2].value:
                            row[1].fill = red_fill
                            row[3].fill = red_fill
                    wb.save(os.path.join(args.output_dir + "/../", 'pred_result.xlsx'))
            acc_str+=f"Epoch {current_epoch} Top 1 Accuracy is {final_top1_action:05.2f}, {final_top1_noun:05.2f}, {final_top1_verb:05.2f}, Top 5 ACT {final_top5_action:05.2f}\n" if args.composition else f"Epoch {current_epoch} Top 1 Accuracy is {final_top1:05.2f}, Top 5 {final_top5:05.2f}\n"
    
    

    total_time = time.time() - start_time
    total_time_str = str(datetime.timedelta(seconds=int(total_time)))
    print('Training time {}'.format(total_time_str))
    if args.slack_api is not None:
        if global_rank == 0 and args.slack_api:
            Token = args.slack_api # 자신의 Token 입력
            job_name=os.environ["SLURM_JOB_NAME"]
            cluster=os.environ["SLURM_SUBMIT_HOST"]
            job_time=total_time_str
            attach_dict = {
            'color' : '#ff0000',
            'author_name' : 'Job Finish',
            'title' : args.vmae_model,
            'text' : args.output_dir + '\n' + cluster,
            }
            attach_list=[attach_dict] 
            contents=f"Job_name:{job_name}\nTraining time is {job_time}, B:{args.batch_size}, P:{n_parameters}\n" if args.composition else f"Job_name:{job_name}\nTraining time is {job_time}, B:{args.batch_size}, P:{n_parameters}\n"
            contents+=acc_str
            notice_message(Token, "#notice-job", contents, attach_list)
    


if __name__ == '__main__':
    opts, ds_init = get_args()
    if opts.output_dir:
        Path(opts.output_dir).mkdir(parents=True, exist_ok=True)
    main(opts, ds_init)